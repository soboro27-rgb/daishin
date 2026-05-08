from fastapi import APIRouter, Request, Depends, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from database import get_db
import models
from auth import require_branch
from config import templates
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

VALID_CATEGORIES = ["PC", "노트북", "태블릿", "모바일", "프린터", "복합기", "기타전산기기"]
VALID_CONDITIONS = ["상", "중", "하"]

router = APIRouter()


def _check(request: Request):
    user = require_branch(request)
    if not user:
        return None, RedirectResponse("/login", status_code=302)
    return user, None


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    applications = (
        db.query(models.Application)
        .filter(models.Application.user_id == user["user_id"])
        .order_by(models.Application.updated_at.desc())
        .all()
    )
    return templates.TemplateResponse(
        "branch/dashboard.html",
        {"request": request, "session": request.session, "applications": applications},
    )


@router.get("/applications/new", response_class=HTMLResponse)
def new_application_page(request: Request):
    user, redir = _check(request)
    if redir:
        return redir
    return templates.TemplateResponse(
        "branch/new_application.html",
        {"request": request, "session": request.session},
    )


@router.post("/applications/new")
async def create_application(request: Request, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    form = await request.form()

    app = models.Application(
        user_id=user["user_id"],
        status="draft",
        title=form.get("title", ""),
        notes=form.get("notes", ""),
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(app)
    db.flush()

    categories = form.getlist("category[]")
    model_names = form.getlist("model_name[]")
    manufacturers = form.getlist("manufacturer[]")
    years = form.getlist("manufacture_year[]")
    quantities = form.getlist("quantity[]")
    conditions = form.getlist("condition[]")
    descriptions = form.getlist("description[]")

    for i, cat in enumerate(categories):
        if not cat:
            continue
        try:
            year = int(years[i]) if i < len(years) and years[i] else None
        except ValueError:
            year = None
        try:
            qty = int(quantities[i]) if i < len(quantities) and quantities[i] else 1
        except ValueError:
            qty = 1

        item = models.AssetItem(
            application_id=app.id,
            category=cat,
            model_name=model_names[i] if i < len(model_names) else "",
            manufacturer=manufacturers[i] if i < len(manufacturers) else "",
            manufacture_year=year,
            quantity=qty,
            condition=conditions[i] if i < len(conditions) else "중",
            description=descriptions[i] if i < len(descriptions) else "",
        )
        db.add(item)

    action = form.get("action", "draft")
    if action == "submit" and categories:
        app.status = "submitted"
        app.submitted_at = datetime.now()

    db.commit()
    return RedirectResponse(f"/branch/applications/{app.id}", status_code=302)


@router.get("/applications/{app_id}", response_class=HTMLResponse)
def application_detail(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.user_id == user["user_id"],
    ).first()

    if not app:
        return RedirectResponse("/branch/dashboard", status_code=302)

    return templates.TemplateResponse(
        "branch/application_detail.html",
        {"request": request, "session": request.session, "app": app},
    )


@router.post("/applications/{app_id}/submit")
def submit_application(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.user_id == user["user_id"],
        models.Application.status == "draft",
    ).first()

    if app:
        app.status = "submitted"
        app.submitted_at = datetime.now()
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/branch/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/confirm-schedule")
def confirm_schedule(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.user_id == user["user_id"],
        models.Application.status == "scheduled",
    ).first()

    if app and app.schedule:
        app.schedule.branch_confirmed = True
        app.schedule.confirmed_at = datetime.now()
        app.status = "schedule_confirmed"
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/branch/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/confirm-price")
def confirm_price(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.user_id == user["user_id"],
        models.Application.status == "priced",
    ).first()

    if app and app.settlement:
        app.settlement.branch_confirmed = True
        app.settlement.branch_confirmed_at = datetime.now()
        app.status = "branch_confirmed"
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/branch/applications/{app_id}", status_code=302)


@router.get("/assets/template")
def download_asset_template(request: Request):
    user, redir = _check(request)
    if redir:
        return redir

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "자산목록"

    header_fill = PatternFill(start_color="C41230", end_color="C41230", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ["카테고리*", "모델명", "제조사", "제조연도", "수량*", "상태", "비고"]
    col_widths = [16, 22, 16, 12, 8, 8, 32]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    example_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    examples = [
        ["PC", "ThinkPad X1 Carbon", "Lenovo", 2020, 2, "중", "배터리 불량"],
        ["노트북", "EliteBook 840 G6", "HP", 2019, 1, "하", "화면 미세 흠집"],
        ["프린터", "LaserJet Pro M404n", "HP", 2021, 3, "상", ""],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill

    ws2 = wb.create_sheet("작성요령")
    ws2["A1"] = "자산목록 작성 요령"
    ws2["A1"].font = Font(bold=True, size=13)
    notes = [
        ("카테고리*", f"필수. 다음 중 하나: {', '.join(VALID_CATEGORIES)}"),
        ("모델명", "장비 모델명 (예: ThinkPad X1 Carbon)"),
        ("제조사", "제조사명 (예: Lenovo, HP, Samsung). 비워도 됨."),
        ("제조연도", "4자리 연도 (예: 2020). 비워도 됨."),
        ("수량*", "필수. 1 이상의 정수. 비우면 1로 처리."),
        ("상태", "상/중/하 중 하나. 비우면 '중' 처리."),
        ("비고", "특이사항 (예: 배터리 불량, 화면 흠집 등). 비워도 됨."),
    ]
    for i, (field, desc) in enumerate(notes, 3):
        ws2[f"A{i}"] = field
        ws2[f"A{i}"].font = Font(bold=True)
        ws2[f"B{i}"] = desc
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 65

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''%EC%9E%90%EC%82%B0%EB%AA%A9%EB%A1%9D%EC%96%91%EC%8B%9D.xlsx"},
    )


@router.post("/assets/parse-excel")
async def parse_asset_excel(request: Request, file: UploadFile = File(...)):
    user, redir = _check(request)
    if redir:
        return JSONResponse({"error": "로그인이 필요합니다."}, status_code=401)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents))
    except Exception:
        return JSONResponse({"error": "올바른 엑셀 파일(.xlsx)이 아닙니다."})

    ws = wb.active
    items = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        category = str(row[0]).strip() if row[0] is not None else ""
        model_name = str(row[1]).strip() if row[1] is not None else ""
        manufacturer = str(row[2]).strip() if row[2] is not None else ""
        manufacture_year_raw = row[3]
        quantity_raw = row[4]
        condition = str(row[5]).strip() if row[5] is not None else ""
        description = str(row[6]).strip() if row[6] is not None else ""

        if not category:
            errors.append(f"{row_num}행: 카테고리가 비어있습니다.")
            continue
        if category not in VALID_CATEGORIES:
            errors.append(f"{row_num}행: 카테고리 '{category}'가 올바르지 않습니다. ({', '.join(VALID_CATEGORIES)} 중 하나여야 합니다)")
            continue

        manufacture_year = None
        if manufacture_year_raw is not None and str(manufacture_year_raw).strip():
            try:
                manufacture_year = int(manufacture_year_raw)
                if not (1990 <= manufacture_year <= 2030):
                    errors.append(f"{row_num}행: 제조연도 {manufacture_year}이 유효하지 않습니다. (1990~2030)")
                    manufacture_year = None
            except (ValueError, TypeError):
                errors.append(f"{row_num}행: 제조연도가 올바른 숫자가 아닙니다.")

        quantity = 1
        if quantity_raw is not None and str(quantity_raw).strip():
            try:
                quantity = int(quantity_raw)
                if quantity < 1:
                    errors.append(f"{row_num}행: 수량은 1 이상이어야 합니다. (1로 처리됨)")
                    quantity = 1
            except (ValueError, TypeError):
                errors.append(f"{row_num}행: 수량이 올바른 숫자가 아닙니다. (1로 처리됨)")

        if condition not in VALID_CONDITIONS:
            condition = "중"

        items.append({
            "category": category,
            "model_name": model_name,
            "manufacturer": manufacturer,
            "manufacture_year": manufacture_year,
            "quantity": quantity,
            "condition": condition,
            "description": description,
        })

    return JSONResponse({"items": items, "errors": errors})
